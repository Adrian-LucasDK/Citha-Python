import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import tkinter as tk
from tkinter import filedialog, messagebox, Text, Scrollbar, Button, Entry, Label
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import numpy as np
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

def analisar_dados(arquivo_csv, pasta_destino):
    """
    Realiza a análise exploratória, visualização de dados e exportação de resultados
    de um arquivo CSV para um arquivo XLSX, incluindo média, mediana, desvio padrão
    e gráficos em planilhas separadas dentro do XLSX. As estatísticas também são
    exibidas em uma interface gráfica, mostrando o número de cestas básicas
    possíveis de comprar por ano.

    Args:
        arquivo_csv (str): O caminho para o arquivo CSV de dados.
        pasta_destino (str): O caminho para a pasta onde o arquivo XLSX será salvo.

    Returns:
        bool: True se a análise for concluída com sucesso, False caso contrário.
    """
    try:
        # 1. Leitura dos Dados
        df = pd.read_csv(arquivo_csv)

        # Verificando se as colunas esperadas existem
        colunas_esperadas = ['ano', 'inflacao', 'salario', 'cesta']
        if not all(col in df.columns for col in colunas_esperadas):
            messagebox.showerror("Erro", f"O arquivo '{arquivo_csv}' não possui as colunas esperadas: {', '.join(colunas_esperadas)}.")
            return False

        # 2. Análise Descritiva
        analise_descritiva = pd.DataFrame() # Inicializar analise_descritiva
        if 'inflacao' in df.columns and pd.api.types.is_numeric_dtype(df['inflacao']):
            analise_descritiva = df[['inflacao', 'salario', 'cesta']].agg(['mean', 'median', 'std']).T
            analise_descritiva = analise_descritiva.rename(columns={'mean': 'Média', 'median': 'Mediana', 'std': 'Desvio Padrão'})

            # Calcular a média de cestas por salário mínimo
            cestas_por_salario_minimo_info = {}
            if 'salario' in df.columns and 'cesta' in df.columns and df['salario'].notna().all() and df['cesta'].notna().all():
                salario_minimo = 1412  # Valor do salário mínimo em 2024 (pode ser ajustado)
                df['cestas_por_salario_minimo_temp'] = (df['salario'] / salario_minimo).round(2)
                media_cestas_por_salario = df['cestas_por_salario_minimo_temp'].mean()
                mediana_cestas_por_salario = df['cestas_por_salario_minimo_temp'].median()
                std_cestas_por_salario = df['cestas_por_salario_minimo_temp'].std()
                cestas_por_salario_minimo_info = {
                    'Média': f"{media_cestas_por_salario:.2f}",
                    'Mediana': f"{mediana_cestas_por_salario:.2f}",
                    'Desvio Padrão': f"{std_cestas_por_salario:.2f}"
                }
                df.drop(columns=['cestas_por_salario_minimo_temp'], inplace=True)

            # Calcular o número de cestas por ano
            if 'ano' in df.columns and 'salario' in df.columns and 'cesta' in df.columns:
                df['cestas_por_ano'] = (df['salario'] / df['cesta']).round(2)
                cestas_por_ano_por_ano = df.groupby('ano')['cestas_por_ano'].mean().reset_index()
                cestas_por_ano_por_ano.rename(columns={'cestas_por_ano': 'Média de Cestas Possíveis'}, inplace=True)

            # Exibir análise descritiva na interface gráfica
            janela_estatisticas = tk.Toplevel()
            janela_estatisticas.title("Análise Descritiva")
            texto_estatisticas = Text(janela_estatisticas, wrap="word", width=80, height=20) # Aumentei width e height
            texto_estatisticas.insert(tk.END, "Estatísticas das Variáveis:\n\n")
            texto_estatisticas.insert(tk.END, analise_descritiva.to_string(float_format='{:.2f}'.format) + "\n\n")

            if cestas_por_salario_minimo_info:
                texto_estatisticas.insert(tk.END, "\nCestas por Salário Mínimo:\n")
                for key, value in cestas_por_salario_minimo_info.items():
                    texto_estatisticas.insert(tk.END, f"{key}: {value}\n")
                texto_estatisticas.insert(tk.END, "\n")

            if 'cestas_por_ano_por_ano' in locals():
                texto_estatisticas.insert(tk.END, "Número Médio de Cestas Básicas Possíveis de Comprar por Ano:\n")
                texto_estatisticas.insert(tk.END, cestas_por_ano_por_ano.to_string(index=False, float_format='{:.2f}'.format))
                texto_estatisticas.insert(tk.END, "\n\n")

            texto_estatisticas.config(state="disabled")
            texto_estatisticas.pack(side="top", fill="both", expand=True, padx=10, pady=10)

            scrollbar_estatisticas = Scrollbar(janela_estatisticas, command=texto_estatisticas.yview)
            scrollbar_estatisticas.pack(side="right", fill="y")
            texto_estatisticas.config(yscrollcommand=scrollbar_estatisticas.set)

            wb = Workbook()

            def fechar_janela_estatisticas():
                janela_estatisticas.destroy()
                gerar_e_salvar_resultados(df, arquivo_csv, pasta_destino, wb, analise_descritiva, cestas_por_salario_minimo_info)

            botao_ok_estatisticas = Button(janela_estatisticas, text="OK", command=fechar_janela_estatisticas)
            botao_ok_estatisticas.pack(pady=10)

        else:
            analise_descritiva = pd.DataFrame(columns=['Média', 'Mediana', 'Desvio Padrão'])
            wb = Workbook()
            gerar_e_salvar_resultados(df, arquivo_csv, pasta_destino, wb, analise_descritiva, {})

    except FileNotFoundError:
        messagebox.showerror("Erro", f"O arquivo '{arquivo_csv}' não foi encontrado.")
        return False
    except pd.errors.EmptyDataError:
        messagebox.showerror("Erro", f"O arquivo '{arquivo_csv}' está vazio.")
        return False
    except pd.errors.ParserError:
        messagebox.showerror("Erro", f"Não foi possível analisar o arquivo '{arquivo_csv}'. Verifique o formato.")
        return False
    except ImportError:
        messagebox.showerror("Erro", "A biblioteca openpyxl não está instalada. Por favor, instale-a: pip install openpyxl")
        return False
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro durante a análise: {e}")
        return False

def enviar_email(arquivo_xlsx_path, remetente, senha, destinatario, assunto, corpo=""):
    try:
        msg = MIMEMultipart()
        msg['From'] = remetente
        msg['To'] = destinatario
        msg['Subject'] = assunto

        msg.attach(MIMEText(corpo, 'plain'))

        with open(arquivo_xlsx_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(arquivo_xlsx_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(arquivo_xlsx_path)}"'
            msg.attach(part)

        # Configurações do servidor SMTP (exemplo para Gmail)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remetente, senha)
        server.sendmail(remetente, destinatario, msg.as_string())
        server.quit()
        messagebox.showinfo("Sucesso", "O arquivo XLSX foi enviado por e-mail com sucesso!")
        return True
    except Exception as e:
        messagebox.showerror("Erro ao enviar e-mail", f"Ocorreu um erro ao enviar o e-mail: {e}")
        return False

def janela_envio_email(arquivo_xlsx_path):
    janela_email = tk.Toplevel()
    janela_email.title("Enviar Relatório por E-mail")

    label_remetente = Label(janela_email, text="E-mail do Remetente:")
    label_remetente.grid(row=0, column=0, padx=5, pady=5, sticky="w")
    entry_remetente = Entry(janela_email)
    entry_remetente.grid(row=0, column=1, padx=5, pady=5)

    label_senha = Label(janela_email, text="Senha do Remetente:")
    label_senha.grid(row=1, column=0, padx=5, pady=5, sticky="w")
    entry_senha = Entry(janela_email, show="*")
    entry_senha.grid(row=1, column=1, padx=5, pady=5)

    label_destinatario = Label(janela_email, text="E-mail do Destinatário:")
    label_destinatario.grid(row=2, column=0, padx=5, pady=5, sticky="w")
    entry_destinatario = Entry(janela_email)
    entry_destinatario.grid(row=2, column=1, padx=5, pady=5)

    label_assunto = Label(janela_email, text="Assunto:")
    label_assunto.grid(row=3, column=0, padx=5, pady=5, sticky="w")
    entry_assunto = Entry(janela_email)
    entry_assunto.grid(row=3, column=1, padx=5, pady=5)

    label_corpo = Label(janela_email, text="Corpo (Opcional):")
    label_corpo.grid(row=4, column=0, padx=5, pady=5, sticky="w")
    text_corpo = Text(janela_email, height=5, width=40)
    text_corpo.grid(row=4, column=1, padx=5, pady=5)

    def enviar_e_fechar():
        remetente = entry_remetente.get()
        senha = entry_senha.get()
        destinatario = entry_destinatario.get()
        assunto = entry_assunto.get()
        corpo = text_corpo.get("1.0", tk.END).strip()

        if not remetente or not senha or not destinatario or not assunto:
            messagebox.showerror("Erro", "Preencha todos os campos obrigatórios.")
            return

        if enviar_email(arquivo_xlsx_path, remetente, senha, destinatario, assunto, corpo):
            janela_email.destroy()
            root.destroy()  # Fecha a janela principal após o envio
            print("Programa finalizado.")

    botao_enviar = Button(janela_email, text="Enviar E-mail e Finalizar", command=enviar_e_fechar)
    botao_enviar.grid(row=5, column=0, columnspan=2, pady=10)

def gerar_e_salvar_resultados(df, arquivo_csv, pasta_destino, wb, analise_descritiva, cestas_por_salario_minimo_info):
    """
    Gera os gráficos em PNG e salva na pasta e no arquivo XLSX em planilhas separadas.
    """
    if df is None or df.empty:
        return

    try:
        # 6. Gerando e Inserindo Gráficos nas Planilhas
        pasta_graficos = os.path.join(pasta_destino, 'graficos')
        if not os.path.exists(pasta_graficos):
            os.makedirs(pasta_graficos)

        def salvar_grafico(nome_grafico, titulo, coluna, y_label=None, eixo_y_format=None, tipo='linha', tipo_grafico='simples'):
            if coluna not in df.columns or not pd.api.types.is_numeric_dtype(df[coluna]):
                return None

            nome_arquivo = os.path.join(pasta_graficos, f'{nome_grafico}.png')

            if tipo_grafico == 'simples':
                plt.figure(figsize=(8, 6))
                if tipo == 'linha':
                    plt.plot(df['ano'], df[coluna], marker='o')
                    plt.ylabel(y_label if y_label else coluna.capitalize())
                    if eixo_y_format == 'percent':
                        plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.0%}'.format(y/100)))
                elif tipo == 'barra':
                    plt.figure(figsize=(8, 6))
                    plt.bar(df['ano'], df[coluna])
                    plt.ylabel(y_label if y_label else coluna.capitalize())
                plt.title(titulo)
                plt.xlabel('Ano')
                plt.grid(True)
                plt.xticks(df['ano'])
                plt.tight_layout()
                plt.savefig(nome_arquivo)
                plt.close()
                return nome_arquivo
            elif tipo_grafico == 'box_hist':
                fig, axes = plt.subplots(1, 2, figsize=(12, 5))
                sns.boxplot(y=df[coluna], ax=axes[0])
                axes[0].set_title(f'Boxplot de {coluna.capitalize()}')
                axes[0].set_ylabel(y_label if y_label else coluna.capitalize())
                if eixo_y_format == 'percent':
                    axes[0].yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.0%}'.format(y/100)))

                sns.histplot(df[coluna], kde=True, ax=axes[1])
                axes[1].set_title(f'Histograma de {coluna.capitalize()}')
                axes[1].set_xlabel(y_label if y_label else coluna.capitalize())
                axes[1].set_ylabel('Frequência')
                if eixo_y_format == 'percent':
                    axes[1].yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.0%}'.format(y/100)))
                plt.tight_layout()
                plt.close(fig)
                plt.savefig(nome_arquivo)
                return nome_arquivo
            elif tipo_grafico == 'histograma':
                plt.figure(figsize=(8, 6))
                sns.histplot(df[coluna], kde=True)
                plt.title(f'Histograma de {coluna.capitalize()}')
                plt.xlabel(y_label if y_label else coluna.capitalize())
                plt.ylabel('Frequência')
                plt.tight_layout()
                if eixo_y_format == 'percent':
                    plt.gca().xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: '{:.0%}'.format(x/100)))
                plt.savefig(nome_arquivo)
                plt.close()
                return nome_arquivo
            elif tipo_grafico == 'boxplot':
                plt.figure(figsize=(8, 6))
                sns.boxplot(y=df[coluna])
                plt.title(f'Boxplot de {coluna.capitalize()}')
                plt.ylabel(y_label if y_label else coluna.capitalize())
                plt.tight_layout()
                if eixo_y_format == 'percent':
                    plt.gca().yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.0%}'.format(y/100)))
                plt.savefig(nome_arquivo)
                plt.close()
                return nome_arquivo
            elif tipo_grafico == 'tres_graficos':
                plt.figure(figsize=(10, 6))
                plt.plot(df['ano'], df['inflacao'], marker='o', label='Inflação (%)')
                plt.plot(df['ano'], df['salario'], marker='x', label='Salário')
                plt.plot(df['ano'], df['cesta'], marker='^', label='Valor da Cesta')
                plt.title('Inflação, Salário e Valor da Cesta ao Longo do Tempo')
                plt.xlabel('Ano')
                plt.ylabel('Valor')
                plt.grid(True)
                plt.xticks(df['ano'])
                plt.legend()
                plt.tight_layout()
                plt.savefig(nome_arquivo)
                plt.close()
                return nome_arquivo
            elif tipo_grafico == 'cestas_por_salario':
                plt.figure(figsize=(8, 6))
                if 'salario' in df.columns and 'cesta' in df.columns and df['salario'].notna().all() and df['cesta'].notna().all():
                    df['cestas_por_salario'] = (df['salario'] / df['cesta']).round(2)
                    plt.bar(df['ano'], df['cestas_por_salario'], color='magenta')
                    plt.title('Número de Cestas Possíveis de Comprar com o Salário')
                    plt.xlabel('Ano')
                    plt.ylabel('Número de Cestas')
                    plt.xticks(df['ano'])
                    plt.tight_layout()
                    plt.savefig(nome_arquivo)
                    plt.close()
                    return nome_arquivo
                else:
                    return None
            plt.close()
            return None

        # Gerar e salvar os gráficos em PNG e inserir no XLSX
        graficos_para_inserir = {}

        graficos_para_inserir['inflacao_linha'] = salvar_grafico('inflacao_linha', 'Inflação ao Longo do Tempo', 'inflacao', y_label='Inflação (%)', eixo_y_format='percent', tipo='linha')
        graficos_para_inserir['inflacao_barras'] = salvar_grafico('inflacao_barras', 'Inflação ao Longo do Tempo', 'inflacao', y_label='Inflação (%)', eixo_y_format='percent', tipo='barra')
        graficos_para_inserir['salario_linha'] = salvar_grafico('salario_linha', 'Salário ao Longo do Tempo', 'salario', y_label='Salário', tipo='linha')
        graficos_para_inserir['salario_barras'] = salvar_grafico('salario_barras', 'Salário ao Longo do Tempo', 'salario', y_label='Salário', tipo='barra')
        graficos_para_inserir['cesta_linha'] = salvar_grafico('cesta_linha', 'Valor da Cesta ao Longo do Tempo', 'cesta', y_label='Valor da Cesta', tipo='linha')
        graficos_para_inserir['cesta_barras'] = salvar_grafico('cesta_barras', 'Valor da Cesta ao Longo do Tempo', 'cesta', y_label='Valor da Cesta', tipo='barra')
        graficos_para_inserir['histograma_inflacao'] = salvar_grafico('histograma_inflacao', 'Histograma da Inflação', 'inflacao', y_label='Inflação (%)', eixo_y_format='percent', tipo_grafico='histograma')
        graficos_para_inserir['boxplot_inflacao'] = salvar_grafico('boxplot_inflacao', 'Boxplot da Inflação', 'inflacao', y_label='Inflação (%)', eixo_y_format='percent', tipo_grafico='boxplot')
        graficos_para_inserir['tres_graficos'] = salvar_grafico('tres_graficos', 'Inflação, Salário e Valor da Cesta', 'inflacao', tipo_grafico='tres_graficos')
        graficos_para_inserir['cestas_por_salario'] = salvar_grafico('cestas_por_salario', 'Número de Cestas Possíveis de Comprar com o Salário', 'cesta', tipo_grafico='cestas_por_salario')

        # 4. Exportação de Resultados (Dados em XLSX)
        nome_base = os.path.splitext(os.path.basename(arquivo_csv))[0]
        nome_arquivo_xlsx = f'{nome_base}_analisado.xlsx'
        caminho_arquivo_xlsx = os.path.join(pasta_destino, nome_arquivo_xlsx)

        ws_dados = wb.get_sheet_by_name("Dados") if "Dados" in wb.sheetnames else wb.create_sheet("Dados")
        ws_estatisticas = wb.get_sheet_by_name("Estatísticas") if "Estatísticas" in wb.sheetnames else wb.create_sheet("Estatísticas")

        # Salvando Dados Originais na Planilha "Dados"
        ws_dados.append(df.columns.tolist())
        for row in df.values:
            ws_dados.append(row.tolist())

        # Ajustar a largura das colunas na planilha "Dados"
        for col_idx, col in enumerate(df.columns, 1):
            ws_dados.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)), 10)

        # Salvando Análise Descritiva na Planilha "Estatísticas"
        ws_estatisticas.append(['Média', 'Mediana', 'Desvio Padrão'])
        if not analise_descritiva.empty:
            for index, row in analise_descritiva.iterrows():
                if index in ['inflacao', 'salario', 'cesta']:
                    ws_estatisticas.append([f"{row['Média']:.2f}", f"{row['Mediana']:.2f}", f"{row['Desvio Padrão']:.2f}"])

        # Adicionando informações sobre cestas por salário mínimo
        if cestas_por_salario_minimo_info:
            ws_estatisticas.append([])  # Adiciona uma linha em branco
            ws_estatisticas.append(['Cestas por Salário Mínimo'])
            ws_estatisticas.append([
                cestas_por_salario_minimo_info['Média'],
                cestas_por_salario_minimo_info['Mediana'],
                cestas_por_salario_minimo_info['Desvio Padrão']
            ])

        # Ajustar a largura das colunas na planilha "Estatísticas"
        colunas_estatisticas = ['Média', 'Mediana', 'Desvio Padrão']
        if cestas_por_salario_minimo_info:
            colunas_estatisticas = ['Média', 'Mediana', 'Desvio Padrão'] # Reset para garantir a ordem correta
        for col_idx, col in enumerate(colunas_estatisticas, 1):
            ws_estatisticas.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col)), 15)

        # Inserindo Gráficos nas Planilhas
        for nome_grafico, caminho_imagem in graficos_para_inserir.items():
            if caminho_imagem:
                try:
                    img = Image(caminho_imagem)
                    ws_grafico = wb.create_sheet(title=nome_grafico.replace('_', ' ').title())
                    ws_grafico.add_image(img, 'A1')
                except Exception as e:
                    print(f"Erro ao inserir gráfico '{nome_grafico}' no XLSX: {e}")

        try:
            wb.save(caminho_arquivo_xlsx)
            messagebox.showinfo("Concluído", f"Análise concluída e resultados exportados para: {caminho_arquivo_xlsx}\nGráficos salvos na pasta: {pasta_graficos}")
            janela_envio_email(caminho_arquivo_xlsx) # Chama a janela de envio após salvar
        except Exception as e:
            messagebox.showerror("Erro ao salvar XLSX", f"Ocorreu um erro ao salvar o arquivo XLSX: {e}")

    except Exception as e:
        messagebox.showerror("Erro ao gerar gráficos ou exportar", f"Ocorreu um erro: {e}")

# 3. Interface Gráfica
root = tk.Tk()
root.title("Análise de Dados e Envio por E-mail")

def selecionar_arquivo():
    """Abre uma janela para o usuário selecionar o arquivo CSV."""
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo CSV",
        filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
    )
    return file_path

def selecionar_pasta():
    """Abre uma janela para o usuário selecionar a pasta de destino."""
    pasta = filedialog.askdirectory(title="Selecione a pasta para salvar os resultados")
    return pasta

def executar_analise():
    """Executa a análise, obtendo o arquivo e a pasta via interface gráfica."""
    arquivo = selecionar_arquivo()
    if arquivo:
        pasta_destino = selecionar_pasta()
        if pasta_destino:
            analisar_dados(arquivo, pasta_destino)
        else:
            messagebox.showinfo("Cancelado", "Nenhuma pasta de destino selecionada.")
    else:
        messagebox.showinfo("Cancelado", "Nenhum arquivo selecionado.")

botao_selecionar = Button(root, text="Selecionar Arquivo CSV", command=executar_analise)
botao_selecionar.pack(pady=10)

root.mainloop()